# encoding:utf-8

from xlutils.copy import copy
from openpyxl import load_workbook
import xlrd, os, docx



'''
allKey = [

    dic = {
    
        key : ...  关键字名称
        row : ...  模板中的行数
        col : ...  模板中的列数
        value : ...  新客户的信息
    }
]

'''
        
#替换关键字
def fill_file(allKey, filename):
    if filename.find(".docx") > -1:
        print("正在替换模板：" + filename)
        filepath = os.getcwd() + '\\temple\\' + filename
        doc = docx.Document(filepath)
        for para in doc.paragraphs:
            #runs = copy.deepcopy(para.runs)
            for run in para.runs:

                #进行替换关键字
                for dic in allKey:
                    if run.text.find(dic['key']) > -1:
                        run.text = run.text.replace(dic['key'], dic['value'])
                        #para.clear()
                        #para.add_run(run.text, run.style)
        for table in doc.tables:
            for row in table.rows:
                for cell in row.cells:
                    for dic in allKey:
                        if cell.text.find(dic['key']) > -1:
                            cell.text = cell.text.replace(dic['key'], dic['value'])
        newfilepath = os.getcwd() + '\\new\\' + filename.replace('docx', 'doc')
        doc.save(newfilepath)
        print("新文件：" + filename + "生成完毕")
    elif (filename.find('.xlsx') > -1 ):
        if filename.find('00') > -1:
            print('客户信息不操作')
        else:
            print("正在替换模板：" + filename)
            filepath = os.getcwd() + '\\temple\\' + filename
            #1.打开excel，建立新的excel
            #workbook = xlrd.open_workbook(filepath,formatting_info=True);
            #nb = copy(workbook)
            #for sheetnum in range(len(workbook.sheets())):
            #    sheet = workbook.sheets()[sheetnum]
            #    for row in range(sheet.nrows):
            #        for col in range(sheet.ncols):
            #            cellText = str(sheet.cell_value(row, col))
            #            for dic in allKey:
            #                if cellText.find(dic['key']) > -1:
            #                    cellText = cellText.replace(dic['key'], dic['value'])
            #                    ns = nb.get_sheet(sheetnum)
            #                    setOutCell(ns, col, row, cellText)
#
            #newfilepath = os.getcwd() + '\\new\\' + filename.replace('xlsx', 'xls')
            #nb.save(newfilepath)

            workbook = load_workbook(filename= filepath)

            #1.获取所有合并单元格
            sheet1 = workbook.worksheets[0]
            merge_all_list = []
            for merge_cell in sheet1.merged_cells:
                r1, r2, c1, c2 = merge_cell.min_row, merge_cell.max_row, merge_cell.min_col, merge_cell.max_col
                merge_all_list.append((r1, r2, c1, c2))
            #2.遍历单元格
            for row in sheet1.rows:
                for cell in row:
                    if hasattr(cell, 'value'):
                        for dic in allKey:
                            cell.value = str(cell.value).replace(dic['key'], dic['value'])

            newfilepath = os.getcwd() + '\\new\\' + filename

            workbook.save(newfilepath)


            print("新文件：" + filename + "生成完毕")
    else:
        print(filename +' 不支持该模板,只支持后缀为docx的模板')

#获取关键字
def read_excel():

    #初始化变量列表
    allKey = []

    ''' 读取信息采集文件,获取关键字 '''
    #首先读取模板  获取关键字命名,并获取关键字位置
    excelName = os.getcwd() + '\\temple\\00客户信息采集.xls'

    workbook = xlrd.open_workbook(excelName)
    sheet1 = workbook.sheets()[0]
    #遍历sheet1
    for row in range(sheet1.nrows):
        for col in range(sheet1.ncols):
            cellText = sheet1.cell_value(row, col)
            if cellText.find('$$') > -1:
                #找到变量后,存为字典
                currentDic = {}
                currentDic['key'] = cellText[2:]
                currentDic['row'] = row
                currentDic['col'] = col

                allKey.append(currentDic)
    return allKey


def read_newCustomerExcel(allKey):
    excelName = os.getcwd() + '\\new\\00客户信息采集.xls'
    workbook = xlrd.open_workbook(excelName)
    sheet1 = workbook.sheets()[0]

    #遍历所有的dic
    for dic in allKey:
        #获取新客户信息
        cell = sheet1.cell(dic['row'], dic['col'])
        ctype = cell.ctype 
        if ctype == 2:  # ctype为2且为浮点
            dic['value'] = str(int(sheet1.cell_value(dic['row'], dic['col'])))
        elif ctype == 3:
            date_value = xlrd.xldate_as_datetime(cell.value, 0)
            dic['value'] = date_value.strftime('%Y/%m/%d')
        else :
            dic['value'] = str(sheet1.cell_value(dic['row'], dic['col']))
    return allKey

#获取模板列表
def templeFiles():
    return os.listdir(os.getcwd() + '\\temple')

def setOutCell(outSheet, col, row, value):
    """ Change cell value without changing formatting. """
    def _getOutCell(outSheet, colIndex, rowIndex):
        """ HACK: Extract the internal xlwt cell representation. """
        row = outSheet._Worksheet__rows.get(rowIndex)
        if not row: return None
 
        cell = row._Row__cells.get(colIndex)
        return cell
 
    # HACK to retain cell style.
    previousCell = _getOutCell(outSheet, col, row)
    # END HACK, PART I
 
    outSheet.write(row, col, value)
 
    # HACK, PART II
    if previousCell:
        newCell = _getOutCell(outSheet, col, row)
        if newCell:
            newCell.xf_idx = previousCell.xf_idx

#进行替换
def autoFill():
    
    allKey = read_excel()
    #1.根据关键字获取新客户的信息
    allKey = read_newCustomerExcel(allKey)
    
    #2.将获取到的客户信息,自动填充到各模板并生成新的文件
    for file in templeFiles():
        fill_file(allKey, file)

    
                  
if __name__ == "__main__":
    #print(templeFiles())
    autoFill()
    #read_docx('你好你好.docx');
#read_docx("C:\\Users\\Administrator\\Desktop\\python_word\\你好你好.docx")